"""Google Sheets API parser."""

from __future__ import annotations

import re

from googleapiclient.discovery import build

from src.exceptions import InvalidSpreadsheetUrlError, ParseError, PermissionDeniedError, WorkbookReadError
from src.models.schema import (
    Alignment,
    Border,
    BorderSide,
    CellData,
    ColumnConfig,
    ConditionalFormatCondition,
    ConditionalFormatStyle,
    ConditionalFormattingRule,
    FontStyle,
    Protection,
    RowConfig,
    SheetData,
    WorkbookData,
)
from src.utils.cell_utils import cell_ref, range_ref
from src.utils.color_utils import gsheet_color_to_hex
from src.utils.logger import get_logger

logger = get_logger(__name__)

_SPREADSHEET_ID_RE = re.compile(r"/spreadsheets/d/([a-zA-Z0-9_-]+)")

# Google Sheets border style mapping
_BORDER_STYLE_MAP = {
    "DOTTED": "dotted",
    "DASHED": "dashed",
    "SOLID": "thin",
    "SOLID_MEDIUM": "medium",
    "SOLID_THICK": "thick",
    "DOUBLE": "double",
    "NONE": None,
}

_HALIGN_MAP = {
    "LEFT": "left",
    "CENTER": "center",
    "RIGHT": "right",
    "GENERAL_HORIZONTAL_ALIGNMENT_UNSPECIFIED": None,
}

_VALIGN_MAP = {
    "TOP": "top",
    "MIDDLE": "center",
    "BOTTOM": "bottom",
}


def extract_spreadsheet_id(url: str) -> str:
    """Extract spreadsheetId from a Google Sheets URL."""
    m = _SPREADSHEET_ID_RE.search(url)
    if not m:
        raise InvalidSpreadsheetUrlError(f"Cannot extract spreadsheet ID from URL: {url}")
    return m.group(1)


def _quote_sheet_name_for_range(name: str) -> str:
    # A1 notation: sheet names with non-alphanumeric/underscore chars must be single-quoted and inner single quotes doubled
    if re.search(r"[^A-Za-z0-9_]", name):
        return "'" + name.replace("'", "''") + "'"
    return name


def _parse_border_side(border_data: dict | None) -> BorderSide | None:
    if not border_data:
        return None
    style_key = border_data.get("style", "NONE")
    style = _BORDER_STYLE_MAP.get(style_key)
    color = gsheet_color_to_hex(border_data.get("color"))
    if style is None and color is None:
        return None
    return BorderSide(style=style, color=color)


def _parse_cell(cell_data: dict, row_idx: int, col_idx: int) -> CellData:
    """Parse a single cell from Google Sheets gridData."""
    # Value
    uev = cell_data.get("userEnteredValue", {})
    ev = cell_data.get("effectiveValue", {})

    value = None
    formula = None
    display_value = None

    if "formulaValue" in uev:
        formula = uev["formulaValue"]
    if "stringValue" in uev:
        value = uev["stringValue"]
    elif "numberValue" in uev:
        value = uev["numberValue"]
    elif "boolValue" in uev:
        value = uev["boolValue"]

    # Effective/display value
    if "stringValue" in ev:
        display_value = ev["stringValue"]
    elif "numberValue" in ev:
        display_value = str(ev["numberValue"])
    elif "boolValue" in ev:
        display_value = str(ev["boolValue"])
    elif "formulaValue" in ev:
        display_value = ev["formulaValue"]

    formatted_value = cell_data.get("formattedValue")
    if formatted_value:
        display_value = formatted_value

    if formula and value is None:
        value = display_value

    # Format
    uef = cell_data.get("userEnteredFormat", {})

    # Font
    font_data = uef.get("textFormat", {})
    font = FontStyle(
        name=font_data.get("fontFamily"),
        size=font_data.get("fontSize"),
        bold=font_data.get("bold", False),
        italic=font_data.get("italic", False),
        underline=font_data.get("underline", False),
        color=gsheet_color_to_hex(font_data.get("foregroundColor")),
    ) if font_data else None

    # Fill
    bg_color = uef.get("backgroundColor")
    fill = gsheet_color_to_hex(bg_color)

    # Alignment
    halign = uef.get("horizontalAlignment")
    valign = uef.get("verticalAlignment")
    wrap = uef.get("wrapStrategy")
    alignment = Alignment(
        horizontal=_HALIGN_MAP.get(halign) if halign else None,
        vertical=_VALIGN_MAP.get(valign) if valign else None,
        wrapText=wrap == "WRAP",
    )

    # Number format
    nf = uef.get("numberFormat", {})
    number_format = nf.get("pattern")

    # Borders
    borders_data = uef.get("borders", {})
    border = Border(
        top=_parse_border_side(borders_data.get("top")),
        bottom=_parse_border_side(borders_data.get("bottom")),
        left=_parse_border_side(borders_data.get("left")),
        right=_parse_border_side(borders_data.get("right")),
    ) if borders_data else None

    return CellData(
        value=value,
        displayValue=display_value,
        formula=formula,
        fill=fill,
        font=font,
        alignment=alignment,
        numberFormat=number_format,
        border=border,
        protection=None,
        hidden=False,
        merged=False,
    )


def _parse_merges(merges: list[dict], sheet_props: dict) -> list[str]:
    """Convert merge metadata to A1-style range strings."""
    result: list[str] = []
    for m in merges:
        sr = m.get("startRowIndex", 0) + 1
        sc = m.get("startColumnIndex", 0) + 1
        er = m.get("endRowIndex", sr)
        ec = m.get("endColumnIndex", sc)
        result.append(range_ref(sr, sc, er, ec))
    return result


def _get_merged_coords(merges: list[dict]) -> set[str]:
    """Build a set of all cell refs that are part of a merge."""
    coords: set[str] = set()
    for m in merges:
        sr = m.get("startRowIndex", 0) + 1
        sc = m.get("startColumnIndex", 0) + 1
        er = m.get("endRowIndex", sr)
        ec = m.get("endColumnIndex", sc)
        for r in range(sr, er + 1):
            for c in range(sc, ec + 1):
                coords.add(cell_ref(r, c))
    return coords


def _parse_conditional_formats(cfs: list[dict]) -> list[ConditionalFormattingRule]:
    """Parse conditional format rules."""
    rules: list[ConditionalFormattingRule] = []
    for idx, cf in enumerate(cfs):
        ranges_data = cf.get("ranges", [])
        range_strs: list[str] = []
        for rd in ranges_data:
            sr = rd.get("startRowIndex", 0) + 1
            sc = rd.get("startColumnIndex", 0) + 1
            er = rd.get("endRowIndex", sr)
            ec = rd.get("endColumnIndex", sc)
            range_strs.append(range_ref(sr, sc, er, ec))

        bool_rule = cf.get("booleanRule", {})
        cond_data = bool_rule.get("condition", {})
        fmt_data = bool_rule.get("format", {})

        cond = None
        if cond_data:
            cond_type = cond_data.get("type", "")
            values = [str(v.get("userEnteredValue", "")) for v in cond_data.get("values", [])]
            cond = ConditionalFormatCondition(operator=cond_type, values=values)

        fmt = None
        if fmt_data:
            text_fmt = fmt_data.get("textFormat", {})
            bg = fmt_data.get("backgroundColor")
            fmt = ConditionalFormatStyle(
                fill=gsheet_color_to_hex(bg),
                fontColor=gsheet_color_to_hex(text_fmt.get("foregroundColor")),
                bold=text_fmt.get("bold"),
            )

        rule_type = cond_data.get("type", "CUSTOM_FORMULA") if cond_data else "CUSTOM_FORMULA"
        rules.append(ConditionalFormattingRule(
            priority=idx + 1,
            ranges=range_strs,
            type=rule_type,
            condition=cond,
            format=fmt,
        ))
    return rules


def _parse_sheet(sheet: dict, sheet_index: int) -> SheetData:
    """Parse one sheet from the API response."""
    props = sheet.get("properties", {})
    title = props.get("title", f"Sheet{sheet_index}")
    hidden = props.get("hidden", False)

    grid_props = props.get("gridProperties", {})
    row_count = grid_props.get("rowCount", 1)
    col_count = grid_props.get("columnCount", 1)

    data_list = sheet.get("data", [])
    grid_data = data_list[0] if data_list else {}

    # Row metadata
    row_metadata = grid_data.get("rowMetadata", [])
    rows: dict[str, RowConfig] = {}
    for r_idx, rm in enumerate(row_metadata):
        rows[str(r_idx + 1)] = RowConfig(
            height=rm.get("pixelSize"),
            hidden=rm.get("hiddenByUser", False) or rm.get("hiddenByFilter", False),
        )

    # Column metadata
    col_metadata = grid_data.get("columnMetadata", [])
    columns: dict[str, ColumnConfig] = {}
    from openpyxl.utils.cell import get_column_letter
    for c_idx, cm in enumerate(col_metadata):
        columns[get_column_letter(c_idx + 1)] = ColumnConfig(
            width=cm.get("pixelSize"),
            hidden=cm.get("hiddenByUser", False) or cm.get("hiddenByFilter", False),
        )

    # Merges
    merges = sheet.get("merges", [])
    merged_ranges = _parse_merges(merges, props)
    merged_coords = _get_merged_coords(merges)

    # Conditional formatting
    cfs = sheet.get("conditionalFormats", [])
    cf_rules = _parse_conditional_formats(cfs)

    # Cells
    cells: dict[str, CellData] = {}
    row_data_list = grid_data.get("rowData", [])
    max_row = len(row_data_list)
    max_col = 0

    for r_idx, row_data in enumerate(row_data_list):
        values = row_data.get("values", [])
        if len(values) > max_col:
            max_col = len(values)
        for c_idx, cell_data in enumerate(values):
            ref = cell_ref(r_idx + 1, c_idx + 1)
            cd = _parse_cell(cell_data, r_idx, c_idx)
            cd.merged = ref in merged_coords
            cd.hidden = (
                rows.get(str(r_idx + 1), RowConfig()).hidden
                or columns.get(get_column_letter(c_idx + 1), ColumnConfig()).hidden
            )
            cells[ref] = cd

    used_range = range_ref(1, 1, max(max_row, 1), max(max_col, 1)) if max_row > 0 else None

    return SheetData(
        name=title,
        index=sheet_index,
        hidden=hidden,
        usedRange=used_range,
        rows=rows,
        columns=columns,
        mergedRanges=merged_ranges,
        conditionalFormattingRules=cf_rules,
        cells=cells,
    )


def parse_google_sheet(spreadsheet_url: str, credentials) -> WorkbookData:
    """Parse a Google Sheet and return WorkbookData.

    Fetches workbook metadata first, then each sheet's grid data individually
    to keep peak memory bounded (avoids OOM on Streamlit Cloud for large workbooks).

    Args:
        spreadsheet_url: Full URL of the Google Sheet.
        credentials: Google OAuth credentials object.
    """
    spreadsheet_id = extract_spreadsheet_id(spreadsheet_url)
    logger.info('Fetching metadata for spreadsheet %s', spreadsheet_id)

    try:
        service = build('sheets', 'v4', credentials=credentials)
        meta = service.spreadsheets().get(
            spreadsheetId=spreadsheet_id,
            includeGridData=False,
        ).execute()
    except Exception as e:
        err_str = str(e).lower()
        if '403' in err_str or 'permission' in err_str:
            raise PermissionDeniedError(f'Permission denied for spreadsheet {spreadsheet_id}: {e}') from e
        raise WorkbookReadError(f'Failed to fetch spreadsheet: {e}') from e

    title = meta.get('properties', {}).get('title', 'Untitled')
    sheet_metas = meta.get('sheets', [])

    sheets: list[SheetData] = []
    total = len(sheet_metas)
    for idx, sheet_meta in enumerate(sheet_metas):
        sheet_title = sheet_meta.get('properties', {}).get('title', f'Sheet{idx}')
        logger.info('Fetching sheet %d/%d: %s', idx + 1, total, sheet_title)
        range_expr = _quote_sheet_name_for_range(sheet_title)
        try:
            sheet_result = service.spreadsheets().get(
                spreadsheetId=spreadsheet_id,
                includeGridData=True,
                ranges=[range_expr],
            ).execute()
        except Exception as e:
            err_str = str(e).lower()
            if '403' in err_str or 'permission' in err_str:
                raise PermissionDeniedError(f'Permission denied for spreadsheet {spreadsheet_id}: {e}') from e
            raise WorkbookReadError(f'Failed to fetch sheet {sheet_title!r}: {e}') from e

        sheet_payloads = sheet_result.get('sheets', [])
        if not sheet_payloads:
            logger.warning('No data returned for sheet %s; skipping', sheet_title)
            continue
        try:
            sheets.append(_parse_sheet(sheet_payloads[0], idx))
        except Exception as e:
            logger.error('Error parsing sheet %r: %s', sheet_title, e)
            raise ParseError(f'Failed to parse sheet {sheet_title!r}: {e}') from e
        finally:
            del sheet_result
            del sheet_payloads

    return WorkbookData(
        sourceType='google_sheets',
        sourceName=title,
        sheets=sheets,
    )
