"""Excel (.xlsx) parser using openpyxl."""

from __future__ import annotations

import datetime
from io import BytesIO
from pathlib import Path
from typing import IO

import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.exceptions import ParseError, UnsupportedFileTypeError, WorkbookReadError
from src.models.schema import (
    Alignment,
    Border,
    BorderSide,
    CellData,
    ColumnConfig,
    ConditionalFormatCondition,
    ConditionalFormatStyle,
    ConditionalFormattingRule,
    FontStyle,
    Protection,
    RowConfig,
    SheetData,
    WorkbookData,
)
from src.utils.cell_utils import cell_ref, range_ref
from src.utils.color_utils import openpyxl_color_to_hex
from src.utils.logger import get_logger

logger = get_logger(__name__)


def _convert_value(value: object) -> str | int | float | bool | None:
    """Convert cell value to a JSON-safe type."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    return str(value)


def _border_side(side: object) -> BorderSide | None:
    if side is None:
        return None
    style = getattr(side, "style", None)
    color = getattr(side, "color", None)
    if style is None and color is None:
        return None
    return BorderSide(
        style=style,
        color=openpyxl_color_to_hex(color) if color else None,
    )


def _parse_conditional_formatting(ws: Worksheet, wb: openpyxl.Workbook) -> list[ConditionalFormattingRule]:
    """Extract conditional formatting rules from worksheet."""
    rules: list[ConditionalFormattingRule] = []
    try:
        for cf_list in ws.conditional_formatting:
            # cf_list.cells gives the actual cell ranges as CellRange objects
            range_parts = [str(r) for r in cf_list.cells]
            if not range_parts:
                # Fallback: sqref attribute
                sqref = getattr(cf_list, "sqref", None)
                if sqref:
                    range_parts = [str(r) for r in sqref.ranges]
            for idx, rule in enumerate(cf_list.rules):
                cond = None
                if hasattr(rule, "operator") and rule.operator:
                    values = []
                    if hasattr(rule, "formula") and rule.formula:
                        values = [str(f) for f in rule.formula]
                    cond = ConditionalFormatCondition(
                        operator=rule.operator,
                        values=values,
                    )

                fmt = None
                dxf = getattr(rule, "dxf", None)
                if dxf:
                    fill_color = None
                    font_color = None
                    bold = None
                    if dxf.fill and dxf.fill.fgColor:
                        fill_color = openpyxl_color_to_hex(dxf.fill.fgColor, wb)
                    if dxf.font:
                        if dxf.font.color:
                            font_color = openpyxl_color_to_hex(dxf.font.color, wb)
                        bold = dxf.font.bold
                    fmt = ConditionalFormatStyle(
                        fill=fill_color,
                        fontColor=font_color,
                        bold=bold,
                    )

                rules.append(ConditionalFormattingRule(
                    priority=getattr(rule, "priority", idx + 1) or (idx + 1),
                    ranges=range_parts,
                    type=getattr(rule, "type", "cellIs") or "cellIs",
                    condition=cond,
                    format=fmt,
                ))
    except Exception as e:
        logger.warning("Failed to parse conditional formatting: %s", e)
    return rules


def _parse_sheet(
    ws_formula: Worksheet,
    ws_display: Worksheet,
    sheet_index: int,
    wb: openpyxl.Workbook,
) -> SheetData:
    """Parse a single worksheet into SheetData."""
    max_row = ws_formula.max_row or 1
    max_col = ws_formula.max_column or 1

    used_range = range_ref(1, 1, max_row, max_col)

    # Row configs
    rows: dict[str, RowConfig] = {}
    for r in range(1, max_row + 1):
        rd = ws_formula.row_dimensions.get(r)
        height = rd.height if rd and rd.height else None
        hidden = rd.hidden if rd else False
        rows[str(r)] = RowConfig(height=height, hidden=bool(hidden))

    # Column configs
    columns: dict[str, ColumnConfig] = {}
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        cd = ws_formula.column_dimensions.get(col_letter)
        width = cd.width if cd and cd.width else None
        hidden = cd.hidden if cd else False
        columns[col_letter] = ColumnConfig(width=width, hidden=bool(hidden))

    # Merged ranges
    merged_ranges = [str(mr) for mr in ws_formula.merged_cells.ranges]

    # Build set of merged cell coordinates for quick lookup
    merged_coords: set[str] = set()
    for mr in ws_formula.merged_cells.ranges:
        for row in range(mr.min_row, mr.max_row + 1):
            for col in range(mr.min_col, mr.max_col + 1):
                merged_coords.add(cell_ref(row, col))

    # Conditional formatting
    cf_rules = _parse_conditional_formatting(ws_formula, wb)

    # All cells
    cells: dict[str, CellData] = {}
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            ref = cell_ref(row, col)
            cell_f = ws_formula.cell(row=row, column=col)
            cell_d = ws_display.cell(row=row, column=col)

            # Value and formula
            value = _convert_value(cell_f.value)
            display_value = _convert_value(cell_d.value)
            formula = None
            if cell_f.data_type == "f" and isinstance(cell_f.value, str) and cell_f.value.startswith("="):
                formula = cell_f.value
                value = display_value  # use cached display value

            # Font
            font = None
            if cell_f.font:
                f = cell_f.font
                font = FontStyle(
                    name=f.name,
                    size=f.size,
                    bold=bool(f.bold),
                    italic=bool(f.italic),
                    underline=bool(f.underline) if isinstance(f.underline, bool) else f.underline is not None and f.underline != "none",
                    color=openpyxl_color_to_hex(f.color, wb),
                )

            # Fill - only extract if there's an actual fill pattern set
            fill_color = None
            if cell_f.fill and cell_f.fill.patternType and cell_f.fill.patternType != "none":
                fg = cell_f.fill.fgColor
                if fg:
                    fill_color = openpyxl_color_to_hex(fg, wb)

            # Alignment
            alignment = None
            if cell_f.alignment:
                a = cell_f.alignment
                alignment = Alignment(
                    horizontal=a.horizontal,
                    vertical=a.vertical,
                    wrapText=bool(a.wrap_text),
                )

            # Border
            border = None
            if cell_f.border:
                b = cell_f.border
                border = Border(
                    top=_border_side(b.top),
                    bottom=_border_side(b.bottom),
                    left=_border_side(b.left),
                    right=_border_side(b.right),
                )

            # Protection
            protection = None
            if cell_f.protection:
                p = cell_f.protection
                protection = Protection(
                    locked=bool(p.locked) if p.locked is not None else True,
                    hidden=bool(p.hidden) if p.hidden is not None else False,
                )

            # Number format
            number_format = cell_f.number_format

            cells[ref] = CellData(
                value=value,
                displayValue=str(display_value) if display_value is not None else None,
                formula=formula,
                fill=fill_color,
                font=font,
                alignment=alignment,
                numberFormat=number_format,
                border=border,
                protection=protection,
                hidden=bool(rows.get(str(row), RowConfig()).hidden or columns.get(get_column_letter(col), ColumnConfig()).hidden),
                merged=ref in merged_coords,
            )

    sheet_state = ws_formula.sheet_state
    hidden = sheet_state != "visible"

    return SheetData(
        name=ws_formula.title,
        index=sheet_index,
        hidden=hidden,
        usedRange=used_range,
        rows=rows,
        columns=columns,
        mergedRanges=merged_ranges,
        conditionalFormattingRules=cf_rules,
        cells=cells,
    )


def parse_excel(source: str | Path | IO[bytes], source_name: str | None = None) -> WorkbookData:
    """Parse an .xlsx file and return WorkbookData.

    Args:
        source: File path, Path object, or file-like BytesIO.
        source_name: Display name for the workbook.
    """
    if isinstance(source, (str, Path)):
        path = Path(source)
        if not path.suffix.lower() == ".xlsx":
            raise UnsupportedFileTypeError(f"Unsupported file type: {path.suffix}. Only .xlsx is supported.")
        source_name = source_name or path.name

    try:
        wb_formula = openpyxl.load_workbook(source, data_only=False)
        # Reset stream position for second load
        if hasattr(source, "seek"):
            source.seek(0)
        wb_display = openpyxl.load_workbook(source, data_only=True)
    except Exception as e:
        raise WorkbookReadError(f"Failed to read workbook: {e}") from e

    source_name = source_name or "unknown.xlsx"

    sheets: list[SheetData] = []
    for idx, sheet_name in enumerate(wb_formula.sheetnames):
        try:
            ws_f = wb_formula[sheet_name]
            ws_d = wb_display[sheet_name]
            sheets.append(_parse_sheet(ws_f, ws_d, idx, wb_formula))
        except Exception as e:
            logger.error("Error parsing sheet '%s': %s", sheet_name, e)
            raise ParseError(f"Failed to parse sheet '{sheet_name}': {e}") from e

    wb_formula.close()
    wb_display.close()

    return WorkbookData(
        sourceType="excel",
        sourceName=source_name,
        sheets=sheets,
    )
