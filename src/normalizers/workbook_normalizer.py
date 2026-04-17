from __future__ import annotations

import re
from typing import Optional

from openpyxl.utils.cell import column_index_from_string, get_column_letter

from src.models.schema import (
    Alignment,
    Border,
    BorderSide,
    CellData,
    ColumnConfig,
    ConditionalFormatCondition,
    ConditionalFormatStyle,
    ConditionalFormattingRule,
    FontStyle,
    Protection,
    RowConfig,
    SheetData,
    WorkbookData,
)

_HEX6 = re.compile(r"^[0-9A-Fa-f]{6}$")
_HEX3 = re.compile(r"^[0-9A-Fa-f]{3}$")
_CELL = re.compile(r"^([A-Za-z]{1,3})(\d+)$")


def _normalize_hex_color(color: Optional[str]) -> Optional[str]:
    if color is None:
        return None
    raw = str(color).strip()
    if not raw:
        return None
    if raw.startswith("#"):
        body = raw[1:]
    else:
        body = raw
    if "," in raw and "." in raw:
        parts = [p.strip() for p in raw.replace("(", " ").replace(")", " ").split(",") if p.strip()]
        try:
            nums = [float(x) for x in parts[:4]]
        except ValueError:
            return None
        r = max(0, min(255, int(round(nums[0] * 255))))
        g = max(0, min(255, int(round(nums[1] * 255))))
        b = max(0, min(255, int(round(nums[2] * 255))))
        return f"#{r:02X}{g:02X}{b:02X}"
    if len(body) == 8 and _HEX6.match(body[2:]):
        body = body[2:]
    elif len(body) == 8 and _HEX6.match(body[:6]):
        body = body[:6]
    if _HEX3.match(body):
        body = "".join(c * 2 for c in body)
    if _HEX6.match(body):
        return f"#{body.upper()}"
    return None


def _normalize_cell_coordinate(addr: str) -> str:
    addr = addr.strip().replace("$", "")
    m = _CELL.match(addr)
    if not m:
        return addr.upper()
    col_raw, row = m.group(1), m.group(2)
    try:
        ci = column_index_from_string(col_raw.upper())
        col = get_column_letter(ci)
    except ValueError:
        col = col_raw.upper()
    return f"{col}{row}"


def _normalize_range_string(rng: str) -> str:
    rng = rng.strip()
    if "!" in rng:
        sheet, rest = rng.split("!", 1)
        q = sheet.strip()
        if q.startswith("'") and q.endswith("'"):
            inner = q[1:-1].replace("''", "'")
            sheet_part = "'" + inner.replace("'", "''") + "'"
        else:
            sheet_part = q
        return f"{sheet_part}!{_normalize_range_string(rest)}"
    if ":" in rng:
        a, b = rng.split(":", 1)
        return f"{_normalize_cell_coordinate(a)}:{_normalize_cell_coordinate(b)}"
    return _normalize_cell_coordinate(rng)


def _normalize_used_range(rng: Optional[str]) -> Optional[str]:
    if rng is None:
        return None
    s = rng.strip()
    if not s:
        return None
    return _normalize_range_string(s)


def _normalize_row_key(key: str) -> str:
    key = str(key).strip()
    if key.isdigit() or (key.startswith("-") and key[1:].isdigit()):
        return str(int(key))
    return key


def _normalize_column_key(key: str) -> str:
    key = str(key).strip()
    if key.isdigit():
        return get_column_letter(int(key))
    try:
        return get_column_letter(column_index_from_string(key.upper()))
    except ValueError:
        return key.upper()


def _font_style(f: Optional[FontStyle]) -> Optional[FontStyle]:
    if f is None:
        return None
    return FontStyle(
        name=f.name,
        size=f.size,
        bold=f.bold,
        italic=f.italic,
        underline=f.underline,
        color=_normalize_hex_color(f.color),
    )


def _alignment(a: Optional[Alignment]) -> Optional[Alignment]:
    if a is None:
        return None
    return Alignment(
        horizontal=a.horizontal,
        vertical=a.vertical,
        wrapText=a.wrapText,
    )


def _border_side(b: Optional[BorderSide]) -> Optional[BorderSide]:
    if b is None:
        return None
    return BorderSide(style=b.style, color=_normalize_hex_color(b.color))


def _border(b: Optional[Border]) -> Optional[Border]:
    if b is None:
        return None
    return Border(
        top=_border_side(b.top),
        bottom=_border_side(b.bottom),
        left=_border_side(b.left),
        right=_border_side(b.right),
    )


def _cell_data(c: CellData, detect_hidden: bool) -> CellData:
    cell_hidden = c.hidden if detect_hidden else False
    return CellData(
        value=c.value,
        displayValue=c.displayValue,
        formula=c.formula,
        fill=_normalize_hex_color(c.fill),
        font=_font_style(c.font),
        alignment=_alignment(c.alignment),
        numberFormat=c.numberFormat,
        border=_border(c.border),
        protection=Protection(locked=c.protection.locked, hidden=c.protection.hidden) if c.protection else None,
        hidden=cell_hidden,
        merged=c.merged,
    )


def _cond_condition(c: Optional[ConditionalFormatCondition]) -> Optional[ConditionalFormatCondition]:
    if c is None:
        return None
    return ConditionalFormatCondition(operator=c.operator, values=list(c.values))


def _cond_style(s: Optional[ConditionalFormatStyle]) -> Optional[ConditionalFormatStyle]:
    if s is None:
        return None
    return ConditionalFormatStyle(
        fill=_normalize_hex_color(s.fill),
        fontColor=_normalize_hex_color(s.fontColor),
        bold=s.bold,
    )


def _cond_rule(r: ConditionalFormattingRule) -> ConditionalFormattingRule:
    return ConditionalFormattingRule(
        priority=r.priority,
        ranges=[_normalize_range_string(x) for x in r.ranges],
        type=r.type,
        condition=_cond_condition(r.condition),
        format=_cond_style(r.format),
    )


def _sheet(sheet: SheetData, detect_hidden: bool) -> SheetData:
    rows: dict[str, RowConfig] = {}
    for k, v in sheet.rows.items():
        nk = _normalize_row_key(k)
        row_hidden = v.hidden if detect_hidden else False
        rows[nk] = RowConfig(height=v.height, hidden=row_hidden)

    columns: dict[str, ColumnConfig] = {}
    for k, v in sheet.columns.items():
        nk = _normalize_column_key(k)
        col_hidden = v.hidden if detect_hidden else False
        columns[nk] = ColumnConfig(width=v.width, hidden=col_hidden)

    cells: dict[str, CellData] = {}
    for ref, c in sheet.cells.items():
        nref = _normalize_cell_coordinate(ref.split("!")[-1])
        cells[nref] = _cell_data(c, detect_hidden)

    sheet_hidden = sheet.hidden if detect_hidden else False
    return SheetData(
        name=sheet.name,
        index=sheet.index,
        hidden=sheet_hidden,
        usedRange=_normalize_used_range(sheet.usedRange),
        rows=rows,
        columns=columns,
        mergedRanges=[_normalize_range_string(m) for m in sheet.mergedRanges],
        conditionalFormattingRules=[_cond_rule(r) for r in sheet.conditionalFormattingRules],
        cells=cells,
    )


def normalize(wb: WorkbookData, detect_hidden: bool = True) -> WorkbookData:
    copied = wb.model_copy(deep=True)
    normalized = WorkbookData(
        sourceType=copied.sourceType,
        sourceName=copied.sourceName,
        sheets=[_sheet(s, detect_hidden) for s in copied.sheets],
    )
    return normalized
