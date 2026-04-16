"""Cell reference and range utility functions."""

from openpyxl.utils.cell import get_column_letter


def col_num_to_letter(n: int) -> str:
    """Convert 1-based column number to letter (1 -> 'A', 27 -> 'AA')."""
    return get_column_letter(n)


def cell_ref(row: int, col: int) -> str:
    """Return A1-style reference from 1-based row and column numbers."""
    return f"{get_column_letter(col)}{row}"


def range_ref(start_row: int, start_col: int, end_row: int, end_col: int) -> str:
    """Return A1-style range string."""
    return f"{cell_ref(start_row, start_col)}:{cell_ref(end_row, end_col)}"
